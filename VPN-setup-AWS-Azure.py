import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_vpn_setup_excel(filename="AWS_Azure_VPN_Setup.xlsx"):
    # Create a new workbook
    wb = Workbook()
    
    # Prerequisites Sheet
    ws_prereq = wb.active
    ws_prereq.title = "Prerequisites"
    
    prereq_data = [
        ["Component", "Requirement", "Notes"],
        ["Azure Subscription", "Active subscription with admin access", "Verify permissions for network resource creation"],
        ["AWS Account", "Active account with admin access", "Verify VPC and VPN creation permissions"],
        ["Azure IP Range", "172.16.0.0/16", "For Azure Virtual Network"],
        ["AWS IP Range", "10.0.0.0/16", "For AWS VPC"],
        ["Azure Region", "Select based on location", "Choose region closest to your primary users"],
        ["AWS Region", "Select based on location", "Should be relatively close to Azure region"],
    ]
    
    for row in prereq_data:
        ws_prereq.append(row)

    # Azure Setup Sheet
    ws_azure = wb.create_sheet("Azure Setup")
    
    azure_data = [
        ["Phase", "Resource", "Configuration", "Estimated Time"],
        ["1", "Resource Group", "Name: RG-AzureAWSVPN", "5 minutes"],
        ["1", "Virtual Network", "Name: AzureVNet\nAddress: 172.16.0.0/16", "10 minutes"],
        ["1", "Subnet", "Name: Subnet-AzureVPN\nAddress: 172.16.1.0/24", "5 minutes"],
        ["1", "Gateway Subnet", "Address: /27 size from VNet space", "5 minutes"],
        ["2", "VPN Gateway", "Name: AzureVPNGateway\nSKU: VpnGw1\nType: Route-based", "45 minutes"],
        ["3", "Local Network Gateway", "Name: AWSLocalNetworkGateway\nIP: AWS VPN Public IP", "10 minutes"],
        ["3", "VPN Connection", "Name: AzureAWSVPNConnection\nType: Site-to-site (IPsec)", "15 minutes"]
    ]
    
    for row in azure_data:
        ws_azure.append(row)

    # AWS Setup Sheet
    ws_aws = wb.create_sheet("AWS Setup")
    
    aws_data = [
        ["Phase", "Resource", "Configuration", "Estimated Time"],
        ["1", "VPC", "Name: AWS-VPC\nAddress: 10.0.0.0/16", "10 minutes"],
        ["1", "Subnet", "Name: Subnet-AWSVPN\nAddress: 10.0.1.0/24", "5 minutes"],
        ["2", "Virtual Private Gateway", "Name: AWS-VPN-VGW", "10 minutes"],
        ["2", "Customer Gateway", "Name: Azure-CGW\nIP: Azure VPN Gateway Public IP", "10 minutes"],
        ["3", "Site-to-Site VPN", "Type: Static\nRouting: Static", "15 minutes"],
        ["3", "Route Table", "Add route for Azure subnet", "5 minutes"]
    ]
    
    for row in aws_data:
        ws_aws.append(row)

    # Testing Sheet
    ws_testing = wb.create_sheet("Testing")
    
    testing_data = [
        ["Test Phase", "Test Case", "Expected Result", "Status"],
        ["Initial Connectivity", "VPN Tunnel Status", "Status should show as 'Connected' in both portals", ""],
        ["Initial Connectivity", "Route Propagation", "Routes should appear in route tables", ""],
        ["Network Testing", "Deploy Test VM in Azure", "VM should be created successfully", ""],
        ["Network Testing", "Deploy EC2 in AWS", "EC2 instance should be created successfully", ""],
        ["Network Testing", "Ping Test Azure to AWS", "Ping should succeed using private IPs", ""],
        ["Network Testing", "Ping Test AWS to Azure", "Ping should succeed using private IPs", ""]
    ]
    
    for row in testing_data:
        ws_testing.append(row)

    # Troubleshooting Sheet
    ws_trouble = wb.create_sheet("Troubleshooting")
    
    trouble_data = [
        ["Issue", "Possible Cause", "Resolution Steps"],
        ["VPN Connection Not Established", "Mismatched shared key", "Verify shared key is identical on both sides"],
        ["VPN Connection Not Established", "Security group/NSG rules", "Check ICMP and required ports are allowed"],
        ["Cannot Ping Across VPN", "Route tables not updated", "Verify route propagation and static routes"],
        ["High Latency", "Region selection", "Verify Azure and AWS regions are geographically close"],
        ["Connection Drops", "Dead Peer Detection (DPD)", "Adjust DPD timeout values"],
        ["BGP Not Working", "ASN mismatch", "Verify ASN numbers match on both sides"]
    ]
    
    for row in trouble_data:
        ws_trouble.append(row)

    # Apply formatting
    for ws in [ws_prereq, ws_azure, ws_aws, ws_testing, ws_trouble]:
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    return f"Excel file created: {filename}"

if __name__ == "__main__":
    result = create_vpn_setup_excel()
    print(result)